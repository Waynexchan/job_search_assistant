from datetime import date, datetime
from hashlib import sha1
from pathlib import Path
import re
import shutil

from openpyxl import Workbook, load_workbook
import pandas as pd

from src.rank_jobs import clean_apply_link, standardise_status

try:
    import config
except ImportError:  # pragma: no cover - config exists in normal project runs.
    config = None


PROJECT_ROOT = Path(__file__).resolve().parent
TRACKER_FILE = PROJECT_ROOT / "tracker" / "applications.xlsx"
HISTORY_FILE = PROJECT_ROOT / "tracker" / "applications_history.xlsx"
STATUS_HISTORY_FILE = PROJECT_ROOT / "tracker" / "status_history.xlsx"
STATS_FILE = PROJECT_ROOT / "tracker" / "tracker_stats.xlsx"
ALL_RANKED_JOBS_FILE = PROJECT_ROOT / "output" / "all_ranked_jobs.xlsx"
RANKED_JOBS_FILE = PROJECT_ROOT / "output" / "ranked_jobs.xlsx"
RAW_JOBS_DIR = PROJECT_ROOT / "jobs" / "raw_jobs"
APPLIED_JOBS_DIR = PROJECT_ROOT / "jobs" / "applied_jobs"
ARCHIVED_JOBS_DIR = PROJECT_ROOT / "jobs" / "archived_jobs"
REJECTED_JOBS_DIR = PROJECT_ROOT / "jobs" / "rejected_jobs"
INVALID_RAW_JOBS_DIR = RAW_JOBS_DIR / "_invalid"
RAW_JOB_MOVEMENT_LOG_FILE = PROJECT_ROOT / "output" / "raw_job_file_movement_log.xlsx"

TRACKER_COLUMNS = [
    "job_id",
    "apply_link",
    "company",
    "job_title",
    "source",
    "record_source",
    "current_status",
    "highest_stage",
    "status_date",
    "apply_date",
    "cv_used",
    "cover_letter_used",
    "notes",
    "last_email_subject",
    "last_email_from",
    "last_email_date",
    "created_date",
    "updated_date",
    "raw_job_file_path",
    "archived_raw_job_path",
]

STATUS_HISTORY_COLUMNS = [
    "job_id",
    "apply_link",
    "company",
    "job_title",
    "status",
    "status_date",
    "changed_by",
    "notes",
]

STATUS_SHORTCUTS = {
    "s": "shortlisted",
    "a": "applied",
    "as": "assessment",
    "oa": "assessment",
    "i": "interview",
    "fi": "final_interview",
    "f": "final_interview",
    "o": "offer",
    "r": "rejected",
    "w": "withdrawn",
    "ni": "not_interested",
    "not interested": "not_interested",
    "skip": "skip",
    "e": "expired",
    "c": "closed",
}

ALLOWED_STATUSES = [
    "new",
    "shortlisted",
    "applied",
    "assessment",
    "interview",
    "final_interview",
    "offer",
    "rejected",
    "withdrawn",
    "not_interested",
    "skip",
    "expired",
    "closed",
]

STATUS_ORDER = {
    "new": 0,
    "shortlisted": 1,
    "applied": 2,
    "assessment": 3,
    "interview": 4,
    "final_interview": 5,
    "offer": 6,
}

TERMINAL_STATUSES = ["offer", "rejected", "withdrawn", "not_interested", "skip", "expired", "closed"]
ARCHIVE_STATUSES = ["withdrawn", "not_interested", "skip"]


def cfg(name, default):
    return getattr(config, name, default) if config else default


def clean_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_id_part(value):
    value = clean_text(value).lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value or "unknown"


def base_job_id(company, job_title):
    return f"{normalize_id_part(company)}-{normalize_id_part(job_title)}"


def make_unique_job_id(company, job_title, existing_job_ids, apply_link=""):
    base_id = base_job_id(company, job_title)
    if base_id not in existing_job_ids:
        return base_id

    hash_source = apply_link or f"{company}|{job_title}"
    suffix = sha1(hash_source.encode("utf-8")).hexdigest()[:6]
    return f"{base_id}-{suffix}"


def normalize_status(status):
    status = clean_text(status).lower()
    status = STATUS_SHORTCUTS.get(status, status)
    status = standardise_status(status)

    if status not in ALLOWED_STATUSES:
        allowed = ", ".join(ALLOWED_STATUSES)
        raise ValueError(f"Status must be one of: {allowed}")

    return status


def stage_rank(status):
    return STATUS_ORDER.get(normalize_status(status), -1)


def infer_highest_stage(current_status):
    current_status = normalize_status(current_status)
    if current_status == "offer":
        return "offer"
    if current_status in ["rejected", "withdrawn"]:
        return "applied"
    if current_status in STATUS_ORDER:
        return current_status
    return "new"


def update_highest_stage(existing_highest_stage, new_status):
    new_status = normalize_status(new_status)
    existing_highest_stage = normalize_status(existing_highest_stage or "new")

    if new_status in ["rejected", "withdrawn"]:
        return existing_highest_stage
    if stage_rank(new_status) > stage_rank(existing_highest_stage):
        return new_status
    return existing_highest_stage


def empty_tracker():
    return pd.DataFrame(columns=TRACKER_COLUMNS)


def empty_status_history():
    return pd.DataFrame(columns=STATUS_HISTORY_COLUMNS)


def normalize_tracker_schema(tracker, default_record_source):
    tracker = tracker.copy()

    old_column_map = {
        "Company": "company",
        "Job Title": "job_title",
        "Apply Link": "apply_link",
        "Link": "apply_link",
        "Platform": "source",
        "Date Applied": "apply_date",
        "Status": "current_status",
        "status": "current_status",
        "狀態（Save / Apply / Skip）": "current_status",
    }
    for old_column, new_column in old_column_map.items():
        if old_column in tracker.columns and new_column not in tracker.columns:
            tracker[new_column] = tracker[old_column]

    for column in TRACKER_COLUMNS:
        if column not in tracker.columns:
            tracker[column] = ""

    tracker = tracker[TRACKER_COLUMNS].fillna("")
    tracker["apply_link"] = tracker["apply_link"].apply(clean_apply_link)
    tracker["current_status"] = tracker["current_status"].apply(normalize_status)
    tracker["highest_stage"] = tracker.apply(
        lambda row: normalize_status(row["highest_stage"]) if clean_text(row["highest_stage"]) else infer_highest_stage(row["current_status"]),
        axis=1,
    )
    tracker["record_source"] = tracker["record_source"].apply(clean_text)
    tracker.loc[tracker["record_source"] == "", "record_source"] = default_record_source

    today = date.today().isoformat()
    for column in ["created_date", "updated_date"]:
        tracker[column] = tracker[column].apply(clean_text)
        tracker.loc[tracker[column] == "", column] = today

    existing_ids = set(clean_text(job_id) for job_id in tracker["job_id"] if clean_text(job_id))
    for index, row in tracker.iterrows():
        if clean_text(row["job_id"]):
            continue
        job_id = make_unique_job_id(row["company"], row["job_title"], existing_ids, row["apply_link"])
        tracker.at[index, "job_id"] = job_id
        existing_ids.add(job_id)

    return tracker[TRACKER_COLUMNS]


def load_tracker_file(path, default_record_source):
    if not path.exists():
        return empty_tracker()
    return normalize_tracker_schema(pd.read_excel(path), default_record_source)


def load_status_history():
    if not STATUS_HISTORY_FILE.exists():
        return empty_status_history()

    history = pd.read_excel(STATUS_HISTORY_FILE)
    for column in STATUS_HISTORY_COLUMNS:
        if column not in history.columns:
            history[column] = ""
    history = history[STATUS_HISTORY_COLUMNS].fillna("")
    history["apply_link"] = history["apply_link"].apply(clean_apply_link)
    history["status"] = history["status"].apply(normalize_status)
    return history


def load_trackers():
    active = load_tracker_file(TRACKER_FILE, "tracker")
    history = load_tracker_file(HISTORY_FILE, "historical_import")
    status_history = load_status_history()
    return active, history, status_history


def lookup_ranked_job(identifier):
    empty_job = {
        "apply_link": "",
        "company": "",
        "job_title": "",
        "location": "",
        "file_name": "",
        "input_type": "",
        "source": "",
        "record_source": "manual_status_update",
        "cv_used": "",
        "cover_letter_used": "",
        "found_in_output": False,
    }

    output_files = [ALL_RANKED_JOBS_FILE, RANKED_JOBS_FILE]
    jobs = []
    for output_file in output_files:
        if not output_file.exists():
            continue
        output_jobs = pd.read_excel(output_file).fillna("")
        output_jobs["matched_output_file"] = output_file.name
        jobs.append(output_jobs)

    if not jobs:
        return empty_job

    jobs = pd.concat(jobs, ignore_index=True)
    jobs["clean_apply_link"] = jobs["apply_link"].apply(clean_apply_link) if "apply_link" in jobs.columns else ""

    identifier = clean_text(identifier)
    cleaned_identifier = clean_apply_link(identifier)
    matches = jobs[jobs["clean_apply_link"].astype(str).str.lower() == cleaned_identifier.lower()]

    if matches.empty and "job_id" in jobs.columns:
        matches = jobs[jobs["job_id"].astype(str).str.lower() == identifier.lower()]

    if matches.empty:
        return empty_job

    job = matches.iloc[0]
    return {
        "apply_link": clean_apply_link(job.get("apply_link", "")),
        "company": job.get("company", ""),
        "job_title": job.get("job_title", ""),
        "location": job.get("location", ""),
        "file_name": job.get("file_name", ""),
        "input_type": job.get("input_type", ""),
        "source": job.get("source", ""),
        "record_source": "manual_status_update",
        "cv_used": job.get("recommended_cv", ""),
        "cover_letter_used": job.get("recommended_cover_letter", ""),
        "found_in_output": True,
    }


def find_existing_row(active, history, identifier):
    identifier = clean_text(identifier)
    cleaned_link = clean_apply_link(identifier)

    for table_name, table in [("active", active), ("history", history)]:
        if cleaned_link:
            link_matches = table["apply_link"].astype(str).str.lower() == cleaned_link.lower()
            if link_matches.any():
                return table_name, link_matches

        id_matches = table["job_id"].astype(str).str.lower() == identifier.lower()
        if id_matches.any():
            return table_name, id_matches

    return "", None


def row_from_details(job_details, status, notes, existing_job_ids, fallback_identifier=""):
    today = date.today().isoformat()
    apply_link = clean_apply_link(job_details.get("apply_link", ""))
    company = clean_text(job_details.get("company", ""))
    job_title = clean_text(job_details.get("job_title", ""))
    existing_apply_date = clean_text(job_details.get("apply_date", ""))

    if not company and not job_title and fallback_identifier and not fallback_identifier.startswith("http"):
        job_id = normalize_id_part(fallback_identifier)
    else:
        job_id = make_unique_job_id(company, job_title, existing_job_ids, apply_link)

    if status in ["expired", "closed"]:
        highest_stage = "applied" if existing_apply_date else "new"
    elif status == "shortlisted":
        highest_stage = "shortlisted"
    elif status == "applied":
        highest_stage = "applied"
    else:
        highest_stage = update_highest_stage("new", status)

    return {
        "job_id": job_id,
        "apply_link": apply_link,
        "company": company,
        "job_title": job_title,
        "source": clean_text(job_details.get("source", "")),
        "record_source": clean_text(job_details.get("record_source", "")) or "manual_status_update",
        "current_status": status,
        "highest_stage": highest_stage,
        "status_date": today,
        "apply_date": today if status == "applied" else existing_apply_date,
        "cv_used": clean_text(job_details.get("cv_used", "")),
        "cover_letter_used": clean_text(job_details.get("cover_letter_used", "")),
        "notes": notes,
        "last_email_subject": "",
        "last_email_from": "",
        "last_email_date": "",
        "created_date": today,
        "updated_date": today,
        "raw_job_file_path": clean_text(job_details.get("raw_job_file_path", "")),
        "archived_raw_job_path": clean_text(job_details.get("archived_raw_job_path", "")),
    }


def append_status_history(status_history, row, status, notes, changed_by="mark_status.py"):
    status_date = date.today().isoformat()
    duplicate_mask = (
        (status_history["job_id"].astype(str) == clean_text(row["job_id"]))
        & (status_history["status"].astype(str) == status)
        & (status_history["status_date"].astype(str) == status_date)
    )
    if duplicate_mask.any():
        return status_history

    new_event = pd.DataFrame(
        [
            {
                "job_id": row["job_id"],
                "apply_link": row["apply_link"],
                "company": row["company"],
                "job_title": row["job_title"],
                "status": status,
                "status_date": status_date,
                "changed_by": changed_by,
                "notes": notes,
            }
        ]
    )
    return pd.concat([status_history, new_event], ignore_index=True)


def excel_value(value):
    if pd.isna(value):
        return ""
    return value


def ensure_workbook(path, columns):
    if path.exists():
        workbook = load_workbook(path)
        worksheet = workbook.active
        if worksheet.max_row == 0:
            worksheet.append(columns)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(columns)

    headers = [worksheet.cell(row=1, column=column).value for column in range(1, worksheet.max_column + 1)]
    for column in columns:
        if column not in headers:
            worksheet.cell(row=1, column=len(headers) + 1).value = column
            headers.append(column)

    return workbook, worksheet, {header: index for index, header in enumerate(headers, start=1)}


def sync_tracker_file(path, tracker):
    """Update existing tracker rows by job_id without rebuilding the workbook."""
    workbook, worksheet, headers = ensure_workbook(path, TRACKER_COLUMNS)
    desired = {
        clean_text(row["job_id"]): row
        for _, row in tracker[TRACKER_COLUMNS].iterrows()
        if clean_text(row["job_id"])
    }
    existing_rows = {}

    for row_index in range(2, worksheet.max_row + 1):
        job_id = clean_text(worksheet.cell(row=row_index, column=headers["job_id"]).value)
        if job_id:
            existing_rows[job_id] = row_index

    for job_id, row in desired.items():
        row_index = existing_rows.get(job_id)
        if row_index is None:
            row_index = worksheet.max_row + 1
        for column in TRACKER_COLUMNS:
            worksheet.cell(row=row_index, column=headers[column]).value = excel_value(row[column])

    for job_id, row_index in sorted(existing_rows.items(), key=lambda item: item[1], reverse=True):
        if job_id not in desired:
            worksheet.delete_rows(row_index, 1)

    workbook.save(path)


def append_status_history_file(path, status_history):
    """Append missing status events without clearing existing history rows."""
    workbook, worksheet, headers = ensure_workbook(path, STATUS_HISTORY_COLUMNS)
    existing_events = set()

    for row_index in range(2, worksheet.max_row + 1):
        existing_events.add(
            (
                clean_text(worksheet.cell(row=row_index, column=headers["job_id"]).value),
                clean_text(worksheet.cell(row=row_index, column=headers["status"]).value),
                clean_text(worksheet.cell(row=row_index, column=headers["status_date"]).value),
            )
        )

    for _, row in status_history[STATUS_HISTORY_COLUMNS].iterrows():
        event_key = (
            clean_text(row["job_id"]),
            clean_text(row["status"]),
            clean_text(row["status_date"]),
        )
        if not event_key[0] or event_key in existing_events:
            continue

        row_index = worksheet.max_row + 1
        for column in STATUS_HISTORY_COLUMNS:
            worksheet.cell(row=row_index, column=headers[column]).value = excel_value(row[column])
        existing_events.add(event_key)

    workbook.save(path)


def backup_tracker_files():
    backup_dir = PROJECT_ROOT / "tracker" / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    for path in [TRACKER_FILE, HISTORY_FILE, STATUS_HISTORY_FILE]:
        if not path.exists():
            continue
        backup_path = backup_dir / f"{path.stem}_{timestamp}{path.suffix}"
        shutil.copy2(path, backup_path)
        print(f"Backup: {backup_path}")


def ensure_raw_job_lifecycle_dirs():
    for path in [RAW_JOBS_DIR, APPLIED_JOBS_DIR, ARCHIVED_JOBS_DIR, REJECTED_JOBS_DIR, INVALID_RAW_JOBS_DIR]:
        path.mkdir(parents=True, exist_ok=True)


def sanitize_filename_part(value, fallback="unknown"):
    value = clean_text(value)
    value = re.sub(r"[<>:\"/\\|?*\x00-\x1f]+", "-", value)
    value = re.sub(r"\s+", " ", value).strip(" .-_")
    return value[:90] or fallback


def unique_target_path(target_dir, filename):
    target = target_dir / filename
    if not target.exists():
        return target
    stem = target.stem
    suffix = target.suffix
    counter = 2
    while True:
        candidate = target_dir / f"{stem}__{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def make_archived_raw_filename(status_date, company, job_title, original_filename):
    original = sanitize_filename_part(Path(original_filename).name, "raw_job.txt")
    if not original.lower().endswith(".txt"):
        original = f"{original}.txt"
    return (
        f"{status_date}__"
        f"{sanitize_filename_part(company)}__"
        f"{sanitize_filename_part(job_title)}__"
        f"{original}"
    )


def path_is_inside(path, parent):
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def existing_path_from_value(value):
    value = clean_text(value)
    if not value:
        return None
    path = Path(value)
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path if path.exists() else None


def find_raw_job_source_path(changed_row, job_details):
    for path_value in [
        job_details.get("archived_raw_job_path", ""),
        changed_row.get("archived_raw_job_path", ""),
        job_details.get("raw_job_file_path", ""),
        changed_row.get("raw_job_file_path", ""),
    ]:
        path = existing_path_from_value(path_value)
        if path:
            return path

    file_name = clean_text(job_details.get("file_name", ""))
    if file_name:
        raw_path = RAW_JOBS_DIR / Path(file_name).name
        if raw_path.exists():
            return raw_path

    return None


def movement_log_row(action, old_path, new_path, company, job_title, status, reason, success, warning=""):
    return {
        "moved_at": datetime.now().isoformat(timespec="seconds"),
        "action": action,
        "old_path": str(old_path) if old_path else "",
        "new_path": str(new_path) if new_path else "",
        "company": clean_text(company),
        "job_title": clean_text(job_title),
        "status": status,
        "reason": reason,
        "success": bool(success),
        "warning": warning,
    }


def append_raw_job_movement_log(row):
    RAW_JOB_MOVEMENT_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    columns = ["moved_at", "action", "old_path", "new_path", "company", "job_title", "status", "reason", "success", "warning"]
    if RAW_JOB_MOVEMENT_LOG_FILE.exists():
        log = pd.read_excel(RAW_JOB_MOVEMENT_LOG_FILE).fillna("")
    else:
        log = pd.DataFrame(columns=columns)
    for column in columns:
        if column not in log.columns:
            log[column] = ""
    log = pd.concat([log[columns], pd.DataFrame([row])], ignore_index=True)
    log.to_excel(RAW_JOB_MOVEMENT_LOG_FILE, index=False)


def get_lifecycle_target(status, source_path):
    if status == "applied" and path_is_inside(source_path, RAW_JOBS_DIR):
        return APPLIED_JOBS_DIR, "applied", "status marked applied"
    if status == "rejected":
        if path_is_inside(source_path, APPLIED_JOBS_DIR):
            return REJECTED_JOBS_DIR, "rejected", "applied job later marked rejected"
        if path_is_inside(source_path, RAW_JOBS_DIR) and bool(cfg("MOVE_REJECTED_RAW_JOBS", False)):
            return REJECTED_JOBS_DIR, "rejected", "MOVE_REJECTED_RAW_JOBS enabled"
    if status in ARCHIVE_STATUSES and bool(cfg("MOVE_ARCHIVED_RAW_JOBS", False)):
        return ARCHIVED_JOBS_DIR, "archived", "MOVE_ARCHIVED_RAW_JOBS enabled"
    return None, "", ""


def move_raw_job_file_after_status_update(changed_row, job_details, status, status_date):
    ensure_raw_job_lifecycle_dirs()
    source_path = find_raw_job_source_path(changed_row, job_details)
    company = changed_row.get("company", "") or job_details.get("company", "")
    job_title = changed_row.get("job_title", "") or job_details.get("job_title", "")

    if not source_path:
        warning = ""
        if status == "applied":
            warning = "No matching raw job .txt file found; tracker status was updated but no file was moved."
            print(f"Warning: {warning}")
        append_raw_job_movement_log(
            movement_log_row("not_found", "", "", company, job_title, status, "no matching source file", False, warning)
        )
        return None

    if source_path.suffix.lower() != ".txt":
        warning = f"Matched source is not a .txt file: {source_path}"
        print(f"Warning: {warning}")
        append_raw_job_movement_log(
            movement_log_row("skipped", source_path, "", company, job_title, status, "not a txt file", False, warning)
        )
        return None

    target_dir, action, reason = get_lifecycle_target(status, source_path)
    if not target_dir:
        append_raw_job_movement_log(
            movement_log_row("skipped", source_path, "", company, job_title, status, "movement disabled or not applicable", True)
        )
        return None

    filename = make_archived_raw_filename(status_date, company, job_title, source_path.name)
    target_path = unique_target_path(target_dir, filename)
    shutil.move(str(source_path), str(target_path))
    append_raw_job_movement_log(
        movement_log_row(action, source_path, target_path, company, job_title, status, reason, True)
    )
    print(f"Moved raw job file: {source_path} -> {target_path}")
    return target_path


def move_terminal_rows(active, history):
    terminal_mask = active["current_status"].isin(TERMINAL_STATUSES)
    terminal_rows = active[terminal_mask].copy()
    active = active[~terminal_mask].copy()
    if not terminal_rows.empty:
        history = pd.concat([history, terminal_rows], ignore_index=True)
    history = history.drop_duplicates(subset=["job_id"], keep="last")
    active = active.drop_duplicates(subset=["job_id"], keep="last")
    return active, history


def save_outputs(active, history, status_history):
    TRACKER_FILE.parent.mkdir(parents=True, exist_ok=True)
    backup_tracker_files()
    sync_tracker_file(TRACKER_FILE, active)
    sync_tracker_file(HISTORY_FILE, history)
    append_status_history_file(STATUS_HISTORY_FILE, status_history)
    write_tracker_stats(active, history)


def update_raw_job_archive_path(active, history, job_id, source_path, target_path):
    if not target_path:
        return active, history
    job_id = clean_text(job_id)
    for table in [active, history]:
        if table.empty or "job_id" not in table.columns:
            continue
        mask = table["job_id"].astype(str) == job_id
        if not mask.any():
            continue
        table.loc[mask, "archived_raw_job_path"] = str(target_path)
        if source_path and "raw_job_file_path" in table.columns:
            blank_raw_path = table.loc[mask, "raw_job_file_path"].astype(str).str.strip().eq("")
            table.loc[mask, "raw_job_file_path"] = table.loc[mask, "raw_job_file_path"].where(~blank_raw_path, str(source_path))
    return active, history


def mark_status(identifier, status, notes, job_details_override=None):
    status = normalize_status(status)
    notes = clean_text(notes)
    today = date.today().isoformat()
    active, history, status_history = load_trackers()
    table_name, matches = find_existing_row(active, history, identifier)
    job_details = job_details_override or lookup_ranked_job(identifier)

    if not table_name:
        allowed_new_statuses = ["expired", "closed", "shortlisted", "applied"]
        if status not in allowed_new_statuses:
            raise ValueError("No matching tracker row found. Add the job first, then mark its status.")
        if not job_details.get("found_in_output"):
            raise ValueError("No matching tracker row found in tracker or ranked output files.")

        existing_ids = set(active["job_id"].astype(str)) | set(history["job_id"].astype(str))
        new_row = row_from_details(job_details, status, notes, existing_ids, identifier)
        if status in TERMINAL_STATUSES:
            history = pd.concat([history, pd.DataFrame([new_row])], ignore_index=True)
        else:
            active = pd.concat([active, pd.DataFrame([new_row])], ignore_index=True)
        changed_row = new_row
        action = "Added"
    else:
        table = active if table_name == "active" else history
        for column in ["apply_link", "company", "job_title", "source", "record_source", "cv_used", "cover_letter_used"]:
            value = clean_text(job_details.get(column, ""))
            if value:
                table.loc[matches, column] = value
        table.loc[matches, "highest_stage"] = table.loc[matches, "highest_stage"].apply(lambda current: update_highest_stage(current, status))
        table.loc[matches, "current_status"] = status
        table.loc[matches, "status_date"] = today
        table.loc[matches, "updated_date"] = today
        table.loc[matches, "notes"] = notes
        if status == "applied":
            blank_apply_date = table.loc[matches, "apply_date"].astype(str).str.strip() == ""
            table.loc[matches, "apply_date"] = table.loc[matches, "apply_date"].where(~blank_apply_date, today)
        changed_row = table.loc[matches].iloc[0].to_dict()
        action = "Updated"

    status_history = append_status_history(status_history, changed_row, status, notes)
    active, history = move_terminal_rows(active, history)
    save_outputs(active, history, status_history)
    source_path_before_move = find_raw_job_source_path(changed_row, job_details)
    moved_path = None
    if status == "applied" or status == "rejected" or status in ARCHIVE_STATUSES:
        moved_path = move_raw_job_file_after_status_update(changed_row, job_details, status, today)
    if moved_path:
        active, history = update_raw_job_archive_path(active, history, changed_row.get("job_id", ""), source_path_before_move, moved_path)
        sync_tracker_file(TRACKER_FILE, active)
        sync_tracker_file(HISTORY_FILE, history)
        write_tracker_stats(active, history)
    return action, status, today


def write_tracker_stats(active, history):
    combined = pd.concat([active, history], ignore_index=True)
    combined = normalize_tracker_schema(combined, "tracker_stats")

    applied_mask = (combined["highest_stage"].apply(stage_rank) >= STATUS_ORDER["applied"]) | combined["current_status"].isin(["applied", "rejected", "offer", "withdrawn"])
    total_applied = int(applied_mask.sum())
    total_assessment = int((combined["highest_stage"].apply(stage_rank) >= STATUS_ORDER["assessment"]).sum())
    total_interview = int((combined["highest_stage"].apply(stage_rank) >= STATUS_ORDER["interview"]).sum())
    total_final_interview = int((combined["highest_stage"].apply(stage_rank) >= STATUS_ORDER["final_interview"]).sum())
    total_offers = int(((combined["current_status"] == "offer") | (combined["highest_stage"] == "offer")).sum())
    total_rejected = int((combined["current_status"] == "rejected").sum())

    metrics = pd.DataFrame(
        [
            {"metric": "Total Applied", "value": total_applied},
            {"metric": "Total Assessment", "value": total_assessment},
            {"metric": "Total Interview", "value": total_interview},
            {"metric": "Total Final Interview", "value": total_final_interview},
            {"metric": "Total Offers", "value": total_offers},
            {"metric": "Total Rejected", "value": total_rejected},
            {"metric": "Interview Rate", "value": total_interview / total_applied if total_applied else 0},
            {"metric": "Assessment Rate", "value": total_assessment / total_applied if total_applied else 0},
            {"metric": "Offer Rate", "value": total_offers / total_applied if total_applied else 0},
        ]
    )
    by_source = combined.groupby("source", dropna=False).size().reset_index(name="applications")
    by_cv = combined.groupby("cv_used", dropna=False).size().reset_index(name="applications")

    with pd.ExcelWriter(STATS_FILE) as writer:
        metrics.to_excel(writer, sheet_name="metrics", index=False)
        by_source.to_excel(writer, sheet_name="applications_by_source", index=False)
        by_cv.to_excel(writer, sheet_name="applications_by_cv", index=False)


def main():
    print("Update application status")
    print("Shortcuts: s=shortlisted, a=applied, as=assessment, i=interview, fi=final_interview, o=offer, r=rejected, w=withdrawn, e=expired, c=closed")
    print()
    identifier = input("Paste Apply Link or Job ID: ").strip()
    status = input("Status: ").strip()
    notes = input("Notes optional: ").strip()
    try:
        action, saved_status, status_date = mark_status(identifier, status, notes)
    except ValueError as error:
        print(f"Error: {error}")
        raise SystemExit(1)
    print()
    print(f"{action} tracker row.")
    print(f"Status: {saved_status}")
    print(f"Status Date: {status_date}")
    print(f"Updated: {TRACKER_FILE}")
    print(f"Updated: {HISTORY_FILE}")
    print(f"Updated: {STATUS_HISTORY_FILE}")
    print(f"Updated: {STATS_FILE}")


if __name__ == "__main__":
    main()
