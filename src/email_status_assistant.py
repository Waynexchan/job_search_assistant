import base64
from datetime import date, datetime, timedelta
from email.utils import parsedate_to_datetime
import html
from pathlib import Path
import re
import shutil

from openpyxl import Workbook, load_workbook
import pandas as pd

from src.rank_jobs import clean_apply_link, standardise_status


PROJECT_ROOT = Path(__file__).resolve().parents[1]
CREDENTIALS_FILE = PROJECT_ROOT / "credentials.json"
TOKEN_FILE = PROJECT_ROOT / "token.json"
TRACKER_FILE = PROJECT_ROOT / "tracker" / "applications.xlsx"
TRACKER_HISTORY_FILE = PROJECT_ROOT / "tracker" / "applications_history.xlsx"
STATUS_HISTORY_FILE = PROJECT_ROOT / "tracker" / "status_history.xlsx"
BACKUP_DIR = PROJECT_ROOT / "tracker" / "backups"
REVIEW_FILE = PROJECT_ROOT / "output" / "email_status_review.xlsx"
DRY_RUN_REVIEW_FILE = PROJECT_ROOT / "output" / "email_status_dry_run_review.xlsx"

GMAIL_READONLY_SCOPE = ["https://www.googleapis.com/auth/gmail.readonly"]

TRACKER_COLUMNS = [
    "job_id",
    "apply_link",
    "company",
    "job_title",
    "source",
    "record_source",
    "current_status",
    "highest_stage",
    "status_date",
    "apply_date",
    "cv_used",
    "cover_letter_used",
    "notes",
    "last_email_subject",
    "last_email_from",
    "last_email_date",
    "created_date",
    "updated_date",
]

STATUS_HISTORY_COLUMNS = [
    "job_id",
    "apply_link",
    "company",
    "job_title",
    "status",
    "status_date",
    "changed_by",
    "notes",
]

STATUS_ORDER = {
    "new": 0,
    "shortlisted": 1,
    "applied": 2,
    "assessment": 3,
    "interview": 4,
    "final_interview": 5,
    "offer": 6,
}

IGNORE_PHRASES = [
    "job alert",
    "new jobs",
    "more new jobs",
    "recommended jobs",
    "similar jobs",
    "jobs you may be interested in",
    "web developer job alert",
    "coursera",
    "% off",
    "discount",
    "wrap-up",
    "newsletter",
    "part time positions",
    "false advertising in job posts",
]

PROMOTIONAL_SUBJECT_PHRASES = [
    "coursera plus",
    "sunday wrap-up",
    "job alert",
    "more new jobs",
    "recommended jobs",
    "new jobs for you",
    "our recommendation:",
    "job recommendation",
    "boost your employability",
]

APPLICATION_STATUS_EVIDENCE_PHRASES = [
    "your application",
    "application update",
    "interview invitation",
    "assessment invitation",
    "unfortunately",
    "regret to inform",
    "not progressing",
    "not moving forward",
]

NOISY_SENDER_DOMAINS = [
    "jobs.cwjobs.co.uk",
    "courses.reed.co.uk",
]

NON_JOB_SENDER_PHRASES = [
    "american express",
    "lloyds bank",
    "interactive brokers",
    "coursera",
    "elsa",
    "glassdoor community",
]

KNOWN_JOB_PLATFORM_OR_ATS_PHRASES = [
    "ashby",
    "bamboohr",
    "greenhouse",
    "indeed",
    "jobvite",
    "lever",
    "linkedin",
    "myworkday",
    "smartrecruiters",
    "talentlink",
    "teamtailor",
    "workable",
    "workday",
]

STRONG_APPLICATION_EVIDENCE = [
    "unfortunately",
    "unsuccessful",
    "not be moving forward",
    "regret to inform",
    "after careful consideration",
    "we have reviewed your application",
    "thank you for your job application",
    "thank you for applying",
    "your application",
    "application update",
    "invite you to interview",
    "online assessment",
    "offer of employment",
]

APPLICATION_ACKNOWLEDGEMENT_PHRASES = [
    "application received",
    "we received your application",
    "we've received your application",
    "thank you for applying",
    "thank you for your application",
    "thank you for your job application",
    "your application has been received",
]

NEGATIVE_ASSESSMENT_CONTEXT_PHRASES = [
    "some positions may require an online assessment",
    "if this applies to you",
    "we will send a separate email",
    "we'll send you a separate email",
    "keep an eye on your inbox for an invite",
    "our process includes assessment",
    "stage 1 is assessment",
    "assessment may be required",
]

STRICT_STATUS_PHRASES = {
    "rejected": [
        "unfortunately",
        "unsuccessful",
        "not be moving forward",
        "after careful consideration",
        "regret to inform",
        "we will not be proceeding",
        "we won't be progressing",
        "not progressing your application",
        "we have reviewed your application",
    ],
    "assessment": [
        "you are invited to complete an assessment",
        "please complete the assessment",
        "complete your online assessment",
        "assessment link",
        "testgorilla invitation",
        "deadline to complete the assessment",
        "your assessment is ready",
        "start your assessment",
        "click here to begin the assessment",
        "assessment invitation",
    ],
    "interview": [
        "invite you to interview",
        "invited to interview",
        "schedule an interview",
        "book an interview",
        "interview invitation",
        "teams interview",
        "microsoft teams interview",
        "shortlisted for interview",
        "next stage interview",
    ],
    "final_interview": [
        "final interview",
        "final round",
        "last stage",
    ],
    "offer": [
        "we are pleased to offer you",
        "pleased to offer you the role",
        "offer of employment",
        "formal offer of employment",
        "conditional offer of employment",
        "we would like to offer you the position",
        "offer letter attached",
    ],
}


def clean_text(value):
    if pd.isna(value):
        return ""
    return html.unescape(str(value)).strip()


def normalize_for_match(value):
    value = clean_text(value).lower()
    value = re.sub(r"[^a-z0-9\s]", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def normalize_company(value):
    """Normalize company names for conservative tracker matching."""
    normalized = normalize_for_match(value)
    words_to_remove = {"ltd", "limited", "plc", "group", "uk", "inc", "llc"}
    words = [word for word in normalized.split() if word not in words_to_remove]
    normalized = " ".join(words)
    compact = normalized.replace(" ", "")

    if compact in ["jdcom", "jd"]:
        return "jdcom"
    if compact == "sumup":
        return "sumup"
    if compact in ["carbon", "carbonunderwriting"]:
        return "carbon"

    return compact


def normalize_job_title(value):
    return normalize_for_match(value)


def company_matches(extracted_company, tracker_company):
    extracted = normalize_company(extracted_company)
    tracker = normalize_company(tracker_company)
    return bool(extracted and tracker and extracted == tracker)


def title_tokens(title):
    generic_tokens = {"role", "position", "job", "vacancy"}
    return [
        token
        for token in normalize_job_title(title).split()
        if len(token) >= 3 and token not in generic_tokens
    ]


def title_match_strength(extracted_title, tracker_title):
    """Return exact, near, partial or none using strict token overlap."""
    extracted = normalize_job_title(extracted_title)
    tracker = normalize_job_title(tracker_title)
    if not extracted or not tracker:
        return "none"
    if extracted == tracker:
        return "exact"

    extracted_tokens = title_tokens(extracted)
    tracker_tokens = title_tokens(tracker)
    if not extracted_tokens or not tracker_tokens:
        return "none"

    overlap = len(set(extracted_tokens) & set(tracker_tokens))
    extracted_ratio = overlap / len(set(extracted_tokens))
    tracker_ratio = overlap / len(set(tracker_tokens))

    if extracted_ratio >= 0.75 and tracker_ratio >= 0.75:
        return "near"
    if extracted_ratio >= 0.5 and tracker_ratio >= 0.5:
        return "partial"
    return "none"


def contains_phrase(text, phrase):
    """Phrase match that tolerates punctuation differences."""
    return normalize_for_match(phrase) in normalize_for_match(text)


def sender_contains_phrase(sender, phrase):
    return contains_phrase(clean_text(sender), phrase)


def is_non_job_sender(sender):
    return any(sender_contains_phrase(sender, phrase) for phrase in NON_JOB_SENDER_PHRASES)


def is_known_job_platform_or_ats(sender):
    return any(sender_contains_phrase(sender, phrase) for phrase in KNOWN_JOB_PLATFORM_OR_ATS_PHRASES)


def has_application_status_evidence(email):
    raw_text = f"{email['subject']} {email.get('snippet', '')} {email['body']}"
    return any(contains_phrase(raw_text, phrase) for phrase in APPLICATION_STATUS_EVIDENCE_PHRASES)


def get_pre_status_ignore_reason(email):
    """Ignore known recommendation/course noise before status phrase detection."""
    subject = email["subject"]
    sender = email.get("sender", "")
    domain = sender_domain(sender)

    for phrase in ["our recommendation:", "recommended jobs", "job recommendation", "boost your employability"]:
        if contains_phrase(subject, phrase):
            return f"ignored because subject contains '{phrase}'"

    if domain in NOISY_SENDER_DOMAINS and not has_application_status_evidence(email):
        return f"ignored because sender domain is '{domain}' without application status evidence"

    return ""


def has_strong_application_evidence(email):
    raw_text = f"{email['subject']} {email.get('snippet', '')} {email['body']}"
    return any(contains_phrase(raw_text, phrase) for phrase in STRONG_APPLICATION_EVIDENCE)


def has_application_acknowledgement_evidence(text):
    return any(contains_phrase(text, phrase) for phrase in APPLICATION_ACKNOWLEDGEMENT_PHRASES)


def has_negative_assessment_context(text):
    return any(contains_phrase(text, phrase) for phrase in NEGATIVE_ASSESSMENT_CONTEXT_PHRASES)


def get_ignore_reason(email):
    """Filter noisy alerts, newsletters and marketing emails before classification."""
    raw_text = f"{email['subject']} {email.get('snippet', '')} {email['body']}"
    subject = email["subject"]
    sender = email.get("sender", "")
    normalized = normalize_for_match(raw_text)

    for phrase in NON_JOB_SENDER_PHRASES:
        if sender_contains_phrase(sender, phrase):
            return f"ignored because sender contains '{phrase}'"

    # Promotional/job-alert ignores only apply when there is no strong evidence
    # that this is a real application status update.
    if has_strong_application_evidence(email):
        return ""

    for phrase in PROMOTIONAL_SUBJECT_PHRASES:
        if contains_phrase(subject, phrase):
            return f"ignored because subject contains '{phrase}'"

    for phrase in IGNORE_PHRASES:
        if contains_phrase(raw_text, phrase):
            return f"ignored because email contains '{phrase}'"

    if re.search(r"\bat .+ in .+ and \d+ more new jobs\b", normalized):
        return "ignored because email looks like a job alert listing"

    return ""


def find_status_phrase(text, status):
    """Return the strict phrase that proves a status classification."""
    if status == "assessment" and has_negative_assessment_context(text):
        return ""

    for phrase in STRICT_STATUS_PHRASES[status]:
        if contains_phrase(text, phrase):
            return phrase
    return ""


def extract_company_job_title_from_subject(subject):
    """Extract company/job title from common application email subject formats."""
    subject = clean_text(subject)

    indeed_match = re.search(r"^indeed application:\s*(?P<job_title>.+)$", subject, flags=re.IGNORECASE)
    if indeed_match:
        return "", clean_text(indeed_match.group("job_title"))

    patterns = [
        (
            r"your application to (?P<job_title>.+?) at (?P<company>.+)$",
            "job_title_company",
        ),
        (
            r"your application at (?P<company>.+?) - (?P<job_title>.+)$",
            "company_job_title",
        ),
        (
            r"application update:\s*(?P<job_title>.+)$",
            "job_title_only",
        ),
        (
            r"application for (?P<job_title>.+?) at (?P<company>.+)$",
            "job_title_company",
        ),
        (
            r"application for (?P<job_title>.+)$",
            "job_title_only",
        ),
        (
            r"(?P<job_title>.+?) at (?P<company>.+)$",
            "job_title_company",
        ),
    ]

    for pattern, _ in patterns:
        match = re.search(pattern, subject, flags=re.IGNORECASE)
        if match:
            company = clean_text(match.groupdict().get("company", ""))
            job_title = clean_text(match.groupdict().get("job_title", ""))
            return company, job_title

    return "", ""


def extract_company_job_title_from_body(body):
    """Extract company/job title from common application update body text."""
    body = clean_text(body)
    patterns = [
        r"we greatly appreciate your interest in (?P<company>.+?) and the time you have invested to apply for (?P<job_title>.+?)(?:\.|\n)",
        r"interest in\s+(?:the\s+)?(?P<job_title>.+?)(?:\s+\([^)]+\))?\s+position\s+at\s+(?P<company>.+?)(?:\.|\n|$)",
        r"interest in\s+(?:the\s+)?(?P<job_title>.+?)\s+role\s+at\s+(?P<company>.+?)(?:\.|\n|$)",
        r"(?P<job_title>[A-Za-z0-9&,\-/ ]+?) position at (?P<company>[A-Za-z0-9&.,'\-/ ]+?)(?:\.|\n|$)",
    ]

    for pattern in patterns:
        match = re.search(pattern, body, flags=re.IGNORECASE | re.DOTALL)
        if match:
            company = clean_text(match.group("company"))
            job_title = clean_text(match.group("job_title"))
            return company, job_title

    return "", ""


def is_requisition_reference(value):
    value = normalize_for_match(value)
    return bool(re.fullmatch(r"[a-z]{2,}\s*\d+", value) or re.fullmatch(r"[a-z0-9]{6,}", value))


def is_indeed_application_acknowledgement(email):
    """Indeed Apply confirmations are acknowledgements, not offers."""
    sender = clean_text(email.get("sender", "")).lower()
    subject = clean_text(email.get("subject", ""))
    return "indeedapply@indeed.com" in sender and subject.lower().startswith("indeed application:")


def has_strong_rejection_evidence(text):
    strong_rejection_phrases = [
        "unfortunately",
        "unsuccessful",
        "not be moving forward",
        "after careful consideration",
        "regret to inform",
        "we will not be proceeding",
        "not progressing your application",
    ]
    return any(contains_phrase(text, phrase) for phrase in strong_rejection_phrases)


def has_true_offer_evidence(text):
    """Avoid classifying generic 'job offer' wording as an actual offer."""
    if find_status_phrase(text, "offer"):
        return find_status_phrase(text, "offer")

    return ""


def stage_rank(status):
    return STATUS_ORDER.get(standardise_status(status), -1)


def update_highest_stage(existing_highest_stage, new_status):
    """Rejected/withdrawn must not erase the highest reached stage."""
    new_status = standardise_status(new_status)
    existing_highest_stage = standardise_status(existing_highest_stage or "new")

    if new_status in ["rejected", "withdrawn"]:
        return existing_highest_stage
    if stage_rank(new_status) > stage_rank(existing_highest_stage):
        return new_status
    return existing_highest_stage


def normalize_tracker_table(tracker, tracker_file_name):
    """Normalize one tracker file and remember which file it came from."""
    tracker = tracker.copy()

    if "status" in tracker.columns and "current_status" not in tracker.columns:
        tracker["current_status"] = tracker["status"]

    for column in TRACKER_COLUMNS:
        if column not in tracker.columns:
            tracker[column] = ""

    tracker = tracker[TRACKER_COLUMNS].fillna("")
    tracker["apply_link"] = tracker["apply_link"].apply(clean_apply_link)
    tracker["current_status"] = tracker["current_status"].apply(standardise_status)
    tracker["highest_stage"] = tracker["highest_stage"].apply(
        lambda status: standardise_status(status) if clean_text(status) else "new"
    )
    tracker["matched_tracker_file"] = tracker_file_name
    return tracker


def load_tracker():
    """Load both active and historical applications for matching."""
    tables = []

    if TRACKER_FILE.exists():
        tables.append(normalize_tracker_table(pd.read_excel(TRACKER_FILE), "applications.xlsx"))

    if TRACKER_HISTORY_FILE.exists():
        tables.append(normalize_tracker_table(pd.read_excel(TRACKER_HISTORY_FILE), "applications_history.xlsx"))

    if not tables:
        tracker = pd.DataFrame(columns=TRACKER_COLUMNS)
        tracker["matched_tracker_file"] = "none"
        return tracker

    return pd.concat(tables, ignore_index=True)


def load_status_history():
    if not STATUS_HISTORY_FILE.exists():
        return pd.DataFrame(columns=STATUS_HISTORY_COLUMNS)

    history = pd.read_excel(STATUS_HISTORY_FILE)
    for column in STATUS_HISTORY_COLUMNS:
        if column not in history.columns:
            history[column] = ""
    return history[STATUS_HISTORY_COLUMNS].fillna("")


def excel_value(value):
    if pd.isna(value):
        return ""
    return value


def ensure_workbook(path, columns):
    if path.exists():
        workbook = load_workbook(path)
        worksheet = workbook.active
        if worksheet.max_row == 0:
            worksheet.append(columns)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(columns)

    headers = [worksheet.cell(row=1, column=column).value for column in range(1, worksheet.max_column + 1)]
    for column in columns:
        if column not in headers:
            worksheet.cell(row=1, column=len(headers) + 1).value = column
            headers.append(column)

    return workbook, worksheet, {header: index for index, header in enumerate(headers, start=1)}


def sync_tracker_file(path, tracker):
    """Update existing tracker rows by job_id without rebuilding the workbook."""
    workbook, worksheet, headers = ensure_workbook(path, TRACKER_COLUMNS)
    desired = {
        clean_text(row["job_id"]): row
        for _, row in tracker[TRACKER_COLUMNS].iterrows()
        if clean_text(row["job_id"])
    }
    existing_rows = {}

    for row_index in range(2, worksheet.max_row + 1):
        job_id = clean_text(worksheet.cell(row=row_index, column=headers["job_id"]).value)
        if job_id:
            existing_rows[job_id] = row_index

    for job_id, row in desired.items():
        row_index = existing_rows.get(job_id)
        if row_index is None:
            row_index = worksheet.max_row + 1
        for column in TRACKER_COLUMNS:
            worksheet.cell(row=row_index, column=headers[column]).value = excel_value(row[column])

    for job_id, row_index in sorted(existing_rows.items(), key=lambda item: item[1], reverse=True):
        if job_id not in desired:
            worksheet.delete_rows(row_index, 1)

    workbook.save(path)


def append_status_history_file(path, history):
    """Append missing status events without clearing existing history rows."""
    workbook, worksheet, headers = ensure_workbook(path, STATUS_HISTORY_COLUMNS)
    existing_events = set()

    for row_index in range(2, worksheet.max_row + 1):
        existing_events.add(
            (
                clean_text(worksheet.cell(row=row_index, column=headers["job_id"]).value),
                clean_text(worksheet.cell(row=row_index, column=headers["status"]).value),
                clean_text(worksheet.cell(row=row_index, column=headers["status_date"]).value),
            )
        )

    for _, row in history[STATUS_HISTORY_COLUMNS].iterrows():
        event_key = (
            clean_text(row["job_id"]),
            clean_text(row["status"]),
            clean_text(row["status_date"]),
        )
        if not event_key[0] or event_key in existing_events:
            continue

        row_index = worksheet.max_row + 1
        for column in STATUS_HISTORY_COLUMNS:
            worksheet.cell(row=row_index, column=headers[column]).value = excel_value(row[column])
        existing_events.add(event_key)

    workbook.save(path)


def save_tracker(tracker):
    TRACKER_FILE.parent.mkdir(parents=True, exist_ok=True)

    terminal_mask = (
        (tracker["matched_tracker_file"] == "applications.xlsx")
        & tracker["current_status"].isin(["rejected", "offer", "withdrawn"])
    )
    tracker.loc[terminal_mask, "matched_tracker_file"] = "applications_history.xlsx"

    active = tracker[tracker["matched_tracker_file"] == "applications.xlsx"].copy()
    historical = tracker[tracker["matched_tracker_file"] == "applications_history.xlsx"].copy()
    sync_tracker_file(TRACKER_FILE, active)
    sync_tracker_file(TRACKER_HISTORY_FILE, historical)


def save_status_history(history):
    STATUS_HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    append_status_history_file(STATUS_HISTORY_FILE, history)


def backup_tracker_files(verbose=True):
    """Create timestamped backups before live tracker writes."""
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_paths = []

    for tracker_file in [TRACKER_FILE, TRACKER_HISTORY_FILE, STATUS_HISTORY_FILE]:
        if not tracker_file.exists():
            continue
        backup_path = BACKUP_DIR / f"{tracker_file.stem}_{timestamp}{tracker_file.suffix}"
        shutil.copy2(tracker_file, backup_path)
        backup_paths.append(backup_path)

    if verbose:
        print("Tracker backups created before writing:")
        for backup_path in backup_paths:
            print(f"Backup: {backup_path}")

    return backup_paths


def authenticate_gmail(verbose=True):
    """Authenticate using Gmail read-only scope only."""
    if verbose:
        print("Step 1/7: checking credentials.json")
        print(f"credentials.json exists: {CREDENTIALS_FILE.exists()}")
        print("Step 2/7: loading token.json")
        print(f"token.json exists: {TOKEN_FILE.exists()}")

    try:
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from googleapiclient.discovery import build
    except ImportError as error:
        raise RuntimeError(
            "Missing Gmail API packages. Install google-api-python-client, "
            "google-auth-httplib2, and google-auth-oauthlib."
        ) from error

    credentials = None
    if TOKEN_FILE.exists():
        if verbose:
            print("Loading saved Gmail token...")
        credentials = Credentials.from_authorized_user_file(str(TOKEN_FILE), GMAIL_READONLY_SCOPE)

    if credentials and credentials.expired and credentials.refresh_token:
        if verbose:
            print("Refreshing expired Gmail token...")
        credentials.refresh(Request())

    if not credentials or not credentials.valid:
        if not CREDENTIALS_FILE.exists():
            raise FileNotFoundError(f"Missing {CREDENTIALS_FILE}")
        if verbose:
            print("Step 3/7: starting OAuth flow")
            print("Please complete Google login in the browser.")
        flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_FILE), GMAIL_READONLY_SCOPE)
        credentials = flow.run_local_server(port=0)
        TOKEN_FILE.write_text(credentials.to_json(), encoding="utf-8")

    if verbose:
        print("Step 4/7: building Gmail service")
    return build("gmail", "v1", credentials=credentials)


def get_header(headers, name):
    for header in headers:
        if header.get("name", "").lower() == name.lower():
            return header.get("value", "")
    return ""


def decode_body_part(part):
    data = part.get("body", {}).get("data", "")
    if not data:
        return ""
    return base64.urlsafe_b64decode(data.encode("utf-8")).decode("utf-8", errors="ignore")


def extract_message_text(payload):
    text = decode_body_part(payload)
    for part in payload.get("parts", []) or []:
        mime_type = part.get("mimeType", "")
        if mime_type in ["text/plain", "text/html"]:
            text += "\n" + decode_body_part(part)
        if part.get("parts"):
            text += "\n" + extract_message_text(part)
    return re.sub(r"<[^>]+>", " ", text)


def parse_email_date(date_header):
    try:
        return parsedate_to_datetime(date_header).date().isoformat()
    except (TypeError, ValueError):
        return ""


def fetch_recent_emails(service, days=30, max_emails=20, verbose=True):
    query = (
        f'in:inbox -from:kitkit07285@gmail.com newer_than:{days}d '
        '(application OR applied OR interview OR assessment OR unfortunately OR '
        'unsuccessful OR rejected OR "not be moving forward" OR "next stage" OR '
        '"thank you for applying" OR "application update")'
    )
    messages = []
    page_token = None

    if verbose:
        print("Step 5/7: searching Gmail")
        print(f"Query: {query}")
        print(f"Max emails for this run: {max_emails}")

    while True:
        response = service.users().messages().list(
            userId="me",
            q=query,
            pageToken=page_token,
            maxResults=min(100, max_emails),
        ).execute()
        messages.extend(response.get("messages", []))
        messages = messages[:max_emails]
        page_token = response.get("nextPageToken")
        if not page_token or len(messages) >= max_emails:
            break

    emails = []
    if verbose:
        print(f"Found {len(messages)} messages. Step 6/7: reading messages")

    for index, message in enumerate(messages, start=1):
        if verbose:
            print(f"Reading message {index}/{len(messages)}")
        full_message = service.users().messages().get(
            userId="me",
            id=message["id"],
            format="full",
        ).execute()
        payload = full_message.get("payload", {})
        headers = payload.get("headers", [])
        subject = get_header(headers, "Subject")
        sender = get_header(headers, "From")
        email_date = parse_email_date(get_header(headers, "Date"))
        body = extract_message_text(payload)
        emails.append(
            {
                "message_id": message["id"],
                "subject": subject,
                "sender": sender,
                "email_date": email_date,
                "snippet": full_message.get("snippet", ""),
                "body": body,
            }
        )

    return emails


def classify_email(email):
    """Classify an email conservatively and keep evidence for review."""
    extracted_company, extracted_job_title = extract_company_job_title_from_subject(email["subject"])
    if not extracted_company or not extracted_job_title:
        body_company, body_job_title = extract_company_job_title_from_body(email["body"])
        extracted_company = extracted_company or body_company
        if body_job_title and (not extracted_job_title or is_requisition_reference(extracted_job_title)):
            extracted_job_title = body_job_title

    text = f"{email['subject']} {email.get('snippet', '')} {email['body']}"
    if is_non_job_sender(email.get("sender", "")):
        return {
            "detected_status": "ignore",
            "ignore_reason": "ignored because sender is an obvious non-job sender",
            "classification_evidence": "",
            "extracted_company": extracted_company,
            "extracted_job_title": extracted_job_title,
        }

    pre_status_ignore_reason = get_pre_status_ignore_reason(email)
    if pre_status_ignore_reason:
        return {
            "detected_status": "ignore",
            "ignore_reason": pre_status_ignore_reason,
            "classification_evidence": "",
            "extracted_company": extracted_company,
            "extracted_job_title": extracted_job_title,
        }

    if contains_phrase(text, "we have reviewed your application") and contains_phrase(text, "unfortunately"):
        return {
            "detected_status": "rejected",
            "ignore_reason": "",
            "classification_evidence": "we have reviewed your application + unfortunately",
            "extracted_company": extracted_company,
            "extracted_job_title": extracted_job_title,
        }

    # Strong application status classification takes priority over ignore.
    offer_phrase = has_true_offer_evidence(text)
    if offer_phrase:
        return {
            "detected_status": "offer",
            "ignore_reason": "",
            "classification_evidence": f"matched phrase: {offer_phrase}",
            "extracted_company": extracted_company,
            "extracted_job_title": extracted_job_title,
        }

    for status in ["final_interview", "interview", "assessment", "rejected"]:
        phrase = find_status_phrase(text, status)
        if phrase:
            return {
                "detected_status": status,
                "ignore_reason": "",
                "classification_evidence": f"matched phrase: {phrase}",
                "extracted_company": extracted_company,
                "extracted_job_title": extracted_job_title,
            }

    if is_indeed_application_acknowledgement(email):
        return {
            "detected_status": "applied_acknowledgement",
            "ignore_reason": "",
            "classification_evidence": "Indeed application confirmation email",
            "extracted_company": extracted_company,
            "extracted_job_title": extracted_job_title,
        }

    if has_application_acknowledgement_evidence(text):
        return {
            "detected_status": "applied_acknowledgement",
            "ignore_reason": "",
            "classification_evidence": "application acknowledgement evidence",
            "extracted_company": extracted_company,
            "extracted_job_title": extracted_job_title,
        }

    ignore_reason = get_ignore_reason(email)

    if ignore_reason:
        return {
            "detected_status": "ignore",
            "ignore_reason": ignore_reason,
            "classification_evidence": "",
            "extracted_company": extracted_company,
            "extracted_job_title": extracted_job_title,
        }

    return {
        "detected_status": "unknown",
        "ignore_reason": "",
        "classification_evidence": "",
        "extracted_company": extracted_company,
        "extracted_job_title": extracted_job_title,
    }


def extract_links(text):
    return re.findall(r"https?://[^\s<>\"]+", text)


def sender_domain(sender):
    match = re.search(r"@([A-Za-z0-9.-]+)", sender)
    return match.group(1).lower() if match else ""


def is_near_title_match(extracted_title, tracker_title):
    return title_match_strength(extracted_title, tracker_title) in ["exact", "near"]


def score_match(email, tracker_row, classification):
    extracted_company = clean_text(classification.get("extracted_company", ""))
    extracted_job_title = clean_text(classification.get("extracted_job_title", ""))

    apply_link = clean_text(tracker_row.get("apply_link", ""))
    if apply_link:
        for link in extract_links(email["body"]):
            if clean_apply_link(link).lower() == apply_link.lower():
                return 100, "apply_link match"

    company_match = company_matches(extracted_company, tracker_row.get("company", ""))
    title_strength = title_match_strength(extracted_job_title, tracker_row.get("job_title", ""))

    if extracted_company:
        if not company_match:
            return 0, "different company; extracted company does not match tracker company"

        if extracted_job_title and title_strength in ["exact", "near"]:
            return 90, "exact company + exact/near job title"

        if extracted_job_title and title_strength == "partial":
            return 80, "exact company + partial job title"

        if extracted_job_title:
            return 60, "company only; manual review only; job title mismatch or ambiguous"

        return 60, "company only; manual review only; low confidence because job title missing"

    if extracted_job_title and title_strength in ["exact", "near"]:
        return 50, "job title only; manual review only; low confidence because company missing"

    if extracted_job_title and title_strength == "partial":
        return 40, "partial job title only; manual review only; low confidence because company missing"

    return 0, "no company and no job title; ignore or manual review"


def find_best_tracker_match(email, tracker, classification):
    best_index = None
    best_confidence = 0
    best_reasons = ""

    for index, row in tracker.iterrows():
        confidence, reasons = score_match(email, row, classification)
        if confidence > best_confidence:
            best_index = index
            best_confidence = confidence
            best_reasons = reasons

    return best_index, best_confidence, best_reasons


def validate_safe_update(classification, confidence, tracker_row=None):
    """Validate that an automatic tracker update is safe."""
    tracker_row = tracker_row if tracker_row is not None else {}
    detected_status = classification["detected_status"]
    extracted_company = clean_text(classification.get("extracted_company", ""))
    extracted_job_title = clean_text(classification.get("extracted_job_title", ""))
    matched_company = clean_text(tracker_row.get("company", ""))
    matched_job_title = clean_text(tracker_row.get("job_title", ""))

    company_match_pass = (
        not extracted_company
        or bool(matched_company and company_matches(extracted_company, matched_company))
    )
    title_strength = title_match_strength(extracted_job_title, matched_job_title)
    job_title_match_pass = (
        not extracted_job_title
        or bool(matched_job_title and title_strength in ["exact", "near"])
    )
    status_can_update = detected_status in [
        "rejected",
        "assessment",
        "interview",
        "final_interview",
        "offer",
        "applied_acknowledgement",
    ]
    safe_to_update = bool(
        status_can_update
        and confidence >= 80
        and company_match_pass
        and job_title_match_pass
    )

    return company_match_pass, job_title_match_pass, safe_to_update


def make_review_row(email, classification, confidence, reasons, tracker_row=None):
    """Create one row for manual or dry-run review."""
    tracker_row = tracker_row if tracker_row is not None else {}
    detected_status = classification["detected_status"]
    matched_tracker_file = clean_text(tracker_row.get("matched_tracker_file", "")) or "none"
    rejected_without_known_source = (
        detected_status == "rejected"
        and matched_tracker_file == "none"
        and not is_known_job_platform_or_ats(email.get("sender", ""))
    )
    company_match_pass, job_title_match_pass, safe_to_update = validate_safe_update(
        classification,
        confidence,
        tracker_row,
    )

    if detected_status == "ignore":
        suggested_action = "ignore"
    elif rejected_without_known_source:
        suggested_action = "manual review"
    elif safe_to_update:
        suggested_action = "update tracker"
    else:
        suggested_action = "manual review"

    return {
        "email_date": email["email_date"],
        "email_from": email["sender"],
        "email_subject": email["subject"],
        "email_snippet": email.get("snippet", ""),
        "detected_status": detected_status,
        "ignore_reason": classification.get("ignore_reason", ""),
        "classification_evidence": classification.get("classification_evidence", ""),
        "extracted_company": classification.get("extracted_company", ""),
        "extracted_job_title": classification.get("extracted_job_title", ""),
        "matched_company": clean_text(tracker_row.get("company", "")),
        "matched_job_title": clean_text(tracker_row.get("job_title", "")),
        "matched_job_id": clean_text(tracker_row.get("job_id", "")),
        "matched_tracker_file": matched_tracker_file,
        "match_confidence": confidence,
        "match_reason": reasons,
        "company_match_pass": company_match_pass,
        "job_title_match_pass": job_title_match_pass,
        "safe_to_update": safe_to_update,
        "suggested_action": suggested_action,
    }


def should_update_tracker(status, confidence, tracker_row, classification):
    """Decide whether an email status is allowed to update tracker safely."""
    _, _, safe_to_update = validate_safe_update(classification, confidence, tracker_row)
    return safe_to_update


def can_live_update_review_row(review_row):
    """Final live-mode guard before tracker mutation."""
    detected_status = clean_text(review_row.get("detected_status", ""))
    match_confidence = review_row.get("match_confidence", 0)
    safe_to_update = review_row.get("safe_to_update") is True
    suggested_action = clean_text(review_row.get("suggested_action", "")).lower()

    if detected_status == "ignore":
        return False
    if detected_status == "applied_acknowledgement" and match_confidence < 80:
        return False
    if match_confidence < 80:
        return False
    if not safe_to_update:
        return False
    if suggested_action == "manual review":
        return False

    return True


def append_status_history(history, tracker_row, status, email):
    status_date = email["email_date"] or date.today().isoformat()
    changed_by = "email_status_assistant"
    notes = f"Email: {email['subject']}"
    duplicate_mask = (
        (history["job_id"].astype(str) == clean_text(tracker_row["job_id"]))
        & (history["status"].astype(str) == status)
        & (history["status_date"].astype(str) == status_date)
        & (history["changed_by"].astype(str) == changed_by)
        & (history["notes"].astype(str) == notes)
    )
    if duplicate_mask.any():
        return history

    new_row = pd.DataFrame(
        [
            {
                "job_id": tracker_row["job_id"],
                "apply_link": tracker_row["apply_link"],
                "company": tracker_row["company"],
                "job_title": tracker_row["job_title"],
                "status": status,
                "status_date": status_date,
                "changed_by": changed_by,
                "notes": notes,
            }
        ]
    )
    return pd.concat([history, new_row], ignore_index=True)


def update_tracker_from_email(tracker, history, row_index, status, email):
    today = date.today().isoformat()
    status_date = email["email_date"] or today
    tracker_status = "applied" if status == "applied_acknowledgement" else status

    tracker.at[row_index, "current_status"] = tracker_status
    tracker.at[row_index, "highest_stage"] = update_highest_stage(
        tracker.at[row_index, "highest_stage"],
        tracker_status,
    )
    tracker.at[row_index, "status_date"] = status_date
    tracker.at[row_index, "last_email_subject"] = email["subject"]
    tracker.at[row_index, "last_email_from"] = email["sender"]
    tracker.at[row_index, "last_email_date"] = status_date
    tracker.at[row_index, "updated_date"] = today

    history = append_status_history(history, tracker.loc[row_index], tracker_status, email)
    return tracker, history


def process_emails(dry_run=True, days=30, max_emails=20, verbose=True):
    if verbose:
        print("Email Status Assistant starting")
        print(f"Dry run: {dry_run}")

    service = authenticate_gmail(verbose=verbose)
    if verbose:
        print("Loading tracker files")
    tracker = load_tracker()
    history = load_status_history()
    emails = fetch_recent_emails(service, days=days, max_emails=max_emails, verbose=verbose)
    review_rows = []
    updated_count = 0
    counts = {
        "ignored": 0,
        "rejected": 0,
        "assessment": 0,
        "interview": 0,
        "final_interview": 0,
        "offer": 0,
        "applied_acknowledgement": 0,
        "tracker_updates_possible": 0,
        "manual_review_required": 0,
    }

    if verbose:
        print("Step 7/7: updating tracker" if not dry_run else "Step 7/7: dry-run matching only, tracker will not be updated")

    for email in emails:
        classification = classify_email(email)
        status = classification["detected_status"]

        if status == "ignore":
            counts["ignored"] += 1
            review_rows.append(make_review_row(email, classification, 0, classification["ignore_reason"]))
            continue

        if status == "unknown":
            continue

        counts[status] += 1
        row_index, confidence, reasons = find_best_tracker_match(email, tracker, classification)
        tracker_row = tracker.loc[row_index] if row_index is not None else None
        review_row = make_review_row(email, classification, confidence, reasons, tracker_row)

        can_update = row_index is not None and can_live_update_review_row(review_row)

        if can_update:
            counts["tracker_updates_possible"] += 1
            if not dry_run:
                tracker, history = update_tracker_from_email(tracker, history, row_index, status, email)
            updated_count += 1
            if dry_run:
                review_rows.append(review_row)
        else:
            counts["manual_review_required"] += 1
            review_rows.append(review_row)

    if not dry_run:
        backup_tracker_files(verbose=verbose)
        save_tracker(tracker)
        save_status_history(history)

    REVIEW_FILE.parent.mkdir(parents=True, exist_ok=True)
    review_file = DRY_RUN_REVIEW_FILE if dry_run else REVIEW_FILE
    pd.DataFrame(review_rows).to_excel(review_file, index=False)

    return {
        "emails_checked": len(emails),
        "tracker_updates": updated_count,
        "manual_review": len(review_rows),
        "review_file": review_file,
        "dry_run": dry_run,
        "review_rows": review_rows,
        "counts": counts,
    }
